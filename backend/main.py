"""
FastAPI main application for PDF Email Extractor.
Handles file uploads, processing, and export endpoints.
"""
import os
import uuid
import asyncio
from datetime import datetime
from pathlib import Path
from typing import Optional
from contextlib import asynccontextmanager

from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from .config import (
    UPLOAD_DIR, EXPORT_DIR, MAX_FILE_SIZE_BYTES, 
    ALLOWED_EXTENSIONS, AUTO_DELETE_AFTER_SECONDS
)
from .pdf_processor import PDFProcessor
from .email_extractor import EmailExtractor
from .validator import EmailValidator
from .utils.helpers import cleanup_old_files


# In-memory storage for results (use Redis/DB in production)
results_store: dict = {}


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Lifecycle manager for cleanup tasks."""
    # Startup: Clean old files
    cleanup_old_files(UPLOAD_DIR, AUTO_DELETE_AFTER_SECONDS)
    cleanup_old_files(EXPORT_DIR, AUTO_DELETE_AFTER_SECONDS)
    yield
    # Shutdown: Final cleanup
    cleanup_old_files(UPLOAD_DIR, 0)


app = FastAPI(
    title="PDF Email Extractor",
    description="Extract and validate email addresses from PDF files",
    version="1.0.0",
    lifespan=lifespan
)

# CORS middleware for frontend access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Configure appropriately for production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ==================== Health Check ====================

@app.get("/health")
async def health_check():
    """Health check endpoint for deployment platforms."""
    return {"status": "healthy", "service": "pdf-email-extractor"}


# ==================== Models ====================

class EmailResult(BaseModel):
    email: str
    confidence: float
    is_valid: bool
    context: Optional[str] = None
    name_hint: Optional[str] = None
    domain: str
    validation_details: dict


class ProcessingResult(BaseModel):
    id: str
    status: str
    filename: str
    total_emails: int
    valid_emails: int
    invalid_emails: int
    emails: list[EmailResult]
    processing_time: float
    created_at: str


class UploadResponse(BaseModel):
    id: str
    message: str
    status: str


# ==================== Helper Functions ====================

def validate_file(file: UploadFile) -> None:
    """Validate uploaded file type and size."""
    # Check extension
    ext = Path(file.filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Only {', '.join(ALLOWED_EXTENSIONS)} files are allowed."
        )


async def save_upload_file(file: UploadFile, destination: Path) -> int:
    """Save uploaded file and return file size."""
    total_size = 0
    with open(destination, "wb") as buffer:
        while chunk := await file.read(8192):
            total_size += len(chunk)
            if total_size > MAX_FILE_SIZE_BYTES:
                # Clean up partial file
                buffer.close()
                destination.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=413,
                    detail=f"File too large. Maximum size is {MAX_FILE_SIZE_BYTES // (1024*1024)} MB."
                )
            buffer.write(chunk)
    return total_size


async def process_pdf_task(job_id: str, file_path: Path, filename: str) -> None:
    """Background task to process PDF and extract emails."""
    start_time = datetime.now()
    
    try:
        results_store[job_id]["status"] = "processing"
        
        # Extract text from PDF with progress tracking
        processor = PDFProcessor()
        
        # Get total page count first
        total_pages = processor.get_page_count(str(file_path))
        results_store[job_id]["total_pages"] = total_pages
        results_store[job_id]["current_page"] = 0
        results_store[job_id]["progress_text"] = f"Processing page 0 of {total_pages}..."
        
        # Run extraction in thread pool to not block event loop
        import concurrent.futures
        
        def extract_with_progress():
            def update_progress(current_page: int, total_pages: int):
                results_store[job_id]["current_page"] = current_page
                results_store[job_id]["total_pages"] = total_pages
                results_store[job_id]["progress_text"] = f"Processing page {current_page} of {total_pages}..."
            return processor.extract_text(str(file_path), progress_callback=update_progress)
        
        loop = asyncio.get_event_loop()
        with concurrent.futures.ThreadPoolExecutor() as pool:
            text_content = await loop.run_in_executor(pool, extract_with_progress)
        
        # Extract emails from text (includes duplicates)
        extractor = EmailExtractor()
        raw_emails = extractor.extract_emails_with_duplicates(text_content)
        
        # Debug: print raw email counts
        print(f"DEBUG: Raw emails extracted: {len(raw_emails)}")
        
        # Track duplicates
        email_counts = {}
        for email_data in raw_emails:
            email = email_data["email"]
            if email in email_counts:
                email_counts[email]["count"] += 1
            else:
                email_counts[email] = {"count": 1, "data": email_data}
        
        # Debug: print counts
        print(f"DEBUG: Unique email addresses: {len(email_counts)}")
        duplicate_count = sum(1 for info in email_counts.values() if info["count"] > 1)
        print(f"DEBUG: Emails with count > 1: {duplicate_count}")
        
        # Separate duplicates (count > 1) from unique emails
        duplicates = []
        unique_emails = []
        
        for email, info in email_counts.items():
            if info["count"] > 1:
                duplicates.append({
                    "email": email,
                    "count": info["count"],
                    "domain": email.split("@")[1] if "@" in email else ""
                })
            unique_emails.append(info["data"])
        
        # Validate and score emails
        validator = EmailValidator()
        validated_emails = []
        valid_count = 0
        invalid_count = 0
        
        for email_data in unique_emails:
            validation_result = validator.validate(email_data["email"])
            
            email_result = EmailResult(
                email=email_data["email"],
                confidence=validation_result["confidence"],
                is_valid=validation_result["is_valid"],
                context=email_data.get("context"),
                name_hint=email_data.get("name_hint"),
                domain=validation_result["domain"],
                validation_details=validation_result["details"]
            )
            validated_emails.append(email_result)
            
            if validation_result["is_valid"]:
                valid_count += 1
            else:
                invalid_count += 1
        
        # Sort by confidence (highest first)
        validated_emails.sort(key=lambda x: x.confidence, reverse=True)
        # Sort duplicates by count (highest first)
        duplicates.sort(key=lambda x: x["count"], reverse=True)
        
        processing_time = (datetime.now() - start_time).total_seconds()
        
        # Update results
        results_store[job_id].update({
            "status": "completed",
            "total_emails": len(validated_emails),
            "valid_emails": valid_count,
            "invalid_emails": invalid_count,
            "emails": [e.model_dump() for e in validated_emails],
            "duplicates": duplicates,
            "processing_time": processing_time
        })
        
    except Exception as e:
        results_store[job_id].update({
            "status": "failed",
            "error": str(e)
        })
    
    finally:
        # Clean up uploaded file
        file_path.unlink(missing_ok=True)


# ==================== API Endpoints ====================

@app.post("/upload", response_model=UploadResponse)
async def upload_pdf(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(..., description="PDF file to process")
):
    """
    Upload a PDF file for email extraction.
    Returns a job ID to track processing status.
    """
    validate_file(file)
    
    # Generate unique job ID
    job_id = str(uuid.uuid4())
    
    # Save file
    file_path = UPLOAD_DIR / f"{job_id}_{file.filename}"
    await save_upload_file(file, file_path)
    
    # Initialize result entry
    results_store[job_id] = {
        "id": job_id,
        "status": "queued",
        "filename": file.filename,
        "total_emails": 0,
        "valid_emails": 0,
        "invalid_emails": 0,
        "emails": [],
        "processing_time": 0,
        "created_at": datetime.now().isoformat()
    }
    
    # Start background processing
    background_tasks.add_task(process_pdf_task, job_id, file_path, file.filename)
    
    return UploadResponse(
        id=job_id,
        message="File uploaded successfully. Processing started.",
        status="queued"
    )


async def process_multiple_pdfs_task(job_id: str, file_paths: list[tuple[Path, str]]) -> None:
    """Background task to process multiple PDFs and combine results."""
    start_time = datetime.now()
    
    try:
        results_store[job_id]["status"] = "processing"
        
        processor = PDFProcessor()
        extractor = EmailExtractor()
        validator = EmailValidator()
        
        # Track all emails with counts for duplicates
        email_counts = {}
        filenames_processed = []
        total_files = len(file_paths)
        
        import concurrent.futures
        loop = asyncio.get_event_loop()
        
        for file_index, (file_path, filename) in enumerate(file_paths, 1):
            try:
                results_store[job_id]["current_file"] = filename
                results_store[job_id]["file_progress"] = f"Processing file {file_index} of {total_files}: {filename}"
                filenames_processed.append(filename)
                
                # Get total page count first
                total_pages = processor.get_page_count(str(file_path))
                results_store[job_id]["total_pages"] = total_pages
                results_store[job_id]["current_page"] = 0
                results_store[job_id]["progress_text"] = f"File {file_index}/{total_files} ({filename}): Processing page 0 of {total_pages}..."
                
                # Run extraction in thread pool to not block event loop
                def extract_with_progress(fpath=str(file_path), fidx=file_index, fname=filename):
                    def update_progress(current_page: int, total_pages: int):
                        results_store[job_id]["current_page"] = current_page
                        results_store[job_id]["total_pages"] = total_pages
                        results_store[job_id]["progress_text"] = f"File {fidx}/{total_files} ({fname}): Processing page {current_page} of {total_pages}..."
                    return processor.extract_text(fpath, progress_callback=update_progress)
                
                with concurrent.futures.ThreadPoolExecutor() as pool:
                    text_content = await loop.run_in_executor(pool, extract_with_progress)
                
                # Extract emails from text (with duplicates)
                raw_emails = extractor.extract_emails_with_duplicates(text_content)
                
                # Track all occurrences
                for email_data in raw_emails:
                    email = email_data["email"]
                    if email in email_counts:
                        email_counts[email]["count"] += 1
                    else:
                        email_counts[email] = {"count": 1, "data": email_data}
                    
            except Exception as e:
                print(f"Error processing {filename}: {e}")
            finally:
                # Clean up uploaded file
                file_path.unlink(missing_ok=True)
        
        # Separate duplicates from unique emails
        duplicates = []
        unique_emails = []
        
        for email, info in email_counts.items():
            if info["count"] > 1:
                duplicates.append({
                    "email": email,
                    "count": info["count"],
                    "domain": email.split("@")[1] if "@" in email else ""
                })
            unique_emails.append(info["data"])
        
        # Validate and score unique emails
        all_validated_emails = []
        for email_data in unique_emails:
            validation_result = validator.validate(email_data["email"])
            
            email_result = EmailResult(
                email=email_data["email"],
                confidence=validation_result["confidence"],
                is_valid=validation_result["is_valid"],
                context=email_data.get("context"),
                name_hint=email_data.get("name_hint"),
                domain=validation_result["domain"],
                validation_details=validation_result["details"]
            )
            all_validated_emails.append(email_result)
        
        # Sort by confidence (highest first)
        all_validated_emails.sort(key=lambda x: x.confidence, reverse=True)
        # Sort duplicates by count (highest first)
        duplicates.sort(key=lambda x: x["count"], reverse=True)
        
        valid_count = sum(1 for e in all_validated_emails if e.is_valid)
        invalid_count = len(all_validated_emails) - valid_count
        
        processing_time = (datetime.now() - start_time).total_seconds()
        
        # Update results
        results_store[job_id].update({
            "status": "completed",
            "total_emails": len(all_validated_emails),
            "valid_emails": valid_count,
            "invalid_emails": invalid_count,
            "emails": [e.model_dump() for e in all_validated_emails],
            "duplicates": duplicates,
            "processing_time": processing_time,
            "files_processed": filenames_processed
        })
        
    except Exception as e:
        results_store[job_id].update({
            "status": "failed",
            "error": str(e)
        })


@app.post("/upload-multiple")
async def upload_multiple_pdfs(
    background_tasks: BackgroundTasks,
    files: list[UploadFile] = File(..., description="Multiple PDF files to process")
):
    """
    Upload multiple PDF files for email extraction.
    Returns a single job ID that combines all results.
    """
    # Generate single job ID for all files
    job_id = str(uuid.uuid4())
    file_paths = []
    filenames = []
    
    for file in files:
        validate_file(file)
        file_path = UPLOAD_DIR / f"{job_id}_{file.filename}"
        await save_upload_file(file, file_path)
        file_paths.append((file_path, file.filename))
        filenames.append(file.filename)
    
    # Initialize combined result entry
    results_store[job_id] = {
        "id": job_id,
        "status": "queued",
        "filename": ", ".join(filenames),
        "total_emails": 0,
        "valid_emails": 0,
        "invalid_emails": 0,
        "emails": [],
        "processing_time": 0,
        "created_at": datetime.now().isoformat(),
        "file_count": len(files)
    }
    
    # Start background processing for all files combined
    background_tasks.add_task(process_multiple_pdfs_task, job_id, file_paths)
    
    return {
        "id": job_id,
        "message": f"{len(files)} files uploaded successfully. Processing started.",
        "status": "queued",
        "files": filenames
    }


@app.get("/results/{job_id}")
async def get_results(job_id: str):
    """
    Get the processing results for a specific job.
    """
    if job_id not in results_store:
        raise HTTPException(status_code=404, detail="Job not found")
    
    return results_store[job_id]


@app.get("/export/{job_id}")
async def export_results(
    job_id: str,
    format: str = Query("csv", enum=["csv", "txt", "xlsx"])
):
    """
    Export extracted emails in the specified format.
    """
    if job_id not in results_store:
        raise HTTPException(status_code=404, detail="Job not found")
    
    result = results_store[job_id]
    
    if result["status"] != "completed":
        raise HTTPException(
            status_code=400, 
            detail=f"Job not completed. Current status: {result['status']}"
        )
    
    emails = result["emails"]
    filename = f"emails_{job_id[:8]}"
    export_path = EXPORT_DIR / f"{filename}.{format}"
    
    if format == "csv":
        import csv
        with open(export_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Email", "Confidence", "Valid", "Domain", "Context", "Name Hint"])
            for email in emails:
                writer.writerow([
                    email["email"],
                    f"{email['confidence']:.1f}%",
                    "Yes" if email["is_valid"] else "No",
                    email["domain"],
                    email.get("context", ""),
                    email.get("name_hint", "")
                ])
    
    elif format == "txt":
        with open(export_path, "w", encoding="utf-8") as f:
            f.write(f"PDF Email Extraction Results\n")
            f.write(f"Generated: {datetime.now().isoformat()}\n")
            f.write(f"Source: {result['filename']}\n")
            f.write(f"{'='*50}\n\n")
            f.write(f"Total: {result['total_emails']} | Valid: {result['valid_emails']} | Invalid: {result['invalid_emails']}\n\n")
            
            f.write("Valid Emails:\n")
            f.write("-" * 30 + "\n")
            for email in emails:
                if email["is_valid"]:
                    f.write(f"{email['email']} (Confidence: {email['confidence']:.1f}%)\n")
            
            f.write("\nInvalid/Suspicious Emails:\n")
            f.write("-" * 30 + "\n")
            for email in emails:
                if not email["is_valid"]:
                    f.write(f"{email['email']} (Confidence: {email['confidence']:.1f}%)\n")
    
    elif format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Extracted Emails"
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            headers = ["Email", "Confidence", "Valid", "Domain", "Context", "Name Hint"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for row, email in enumerate(emails, 2):
                ws.cell(row=row, column=1, value=email["email"])
                ws.cell(row=row, column=2, value=f"{email['confidence']:.1f}%")
                ws.cell(row=row, column=3, value="Yes" if email["is_valid"] else "No")
                ws.cell(row=row, column=4, value=email["domain"])
                ws.cell(row=row, column=5, value=email.get("context", ""))
                ws.cell(row=row, column=6, value=email.get("name_hint", ""))
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
            
            wb.save(export_path)
            
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="XLSX export requires openpyxl. Install with: pip install openpyxl"
            )
    
    return FileResponse(
        path=export_path,
        filename=f"{filename}.{format}",
        media_type="application/octet-stream"
    )


@app.get("/status")
async def get_status():
    """Get API status and statistics."""
    total_jobs = len(results_store)
    completed_jobs = sum(1 for r in results_store.values() if r["status"] == "completed")
    
    return {
        "status": "online",
        "version": "1.0.0",
        "total_jobs": total_jobs,
        "completed_jobs": completed_jobs,
        "pending_jobs": total_jobs - completed_jobs
    }


@app.delete("/results/{job_id}")
async def delete_results(job_id: str):
    """Delete results for a specific job."""
    if job_id not in results_store:
        raise HTTPException(status_code=404, detail="Job not found")
    
    del results_store[job_id]
    return {"message": "Results deleted successfully"}


# Serve frontend static files
frontend_path = Path(__file__).resolve().parent.parent / "frontend"
if frontend_path.exists():
    app.mount("/", StaticFiles(directory=str(frontend_path), html=True), name="frontend")
